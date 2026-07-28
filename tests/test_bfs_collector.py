from io import BytesIO

import openpyxl
import pytest
import respx
from httpx import Response

from ch_at_a_glance.collectors.bfs import (
    _CO2_EMISSIONS_CODE,
    _CPI_INDEX_CSV_URL,
    _GDP_GROWTH_CODE,
    _HOSPITAL_BEDS_CODE,
    _JOB_VACANCIES_CODE,
    _LIFE_EXPECTANCY_CODE,
    _POPULATION_COMPONENTS_CODE,
    _UNEMPLOYMENT_CODE,
    _WAGE_INDEX_CODE,
    _resolve_asset_url,
    fetch_co2_emissions,
    fetch_cpi_inflation,
    fetch_gdp_growth,
    fetch_hospital_beds,
    fetch_job_vacancies,
    fetch_life_expectancy,
    fetch_net_migration,
    fetch_real_wages,
    fetch_unemployment_rate,
)


def _mock_resolved_asset(code: str, asset_id: str, content: bytes) -> None:
    """Mock both hops of the two-step resolution: the stable code's landing
    page (which is scraped for whichever numeric asset it currently points
    to) and that asset's actual data."""
    respx.get(f"https://www.bfs.admin.ch/asset/de/{code}").mock(
        return_value=Response(
            200,
            text=f'<a href="https://dam-api.bfs.admin.ch/hub/api/dam/assets/{asset_id}/master">x</a>',
        )
    )
    respx.get(f"https://dam-api.bfs.admin.ch/hub/api/dam/assets/{asset_id}/master").mock(
        return_value=Response(200, content=content)
    )


_WAGE_CSV = (
    '"YEAR","SECTION","DIVISION","SEX","WAGE_TYPE","VALUE","VALUE_P","OBS_STATUS_VALUE"\n'
    '"2023","B-S","05 - 96","T","N","105.9","1.7","A"\n'
    '"2023","B-S","05 - 96","T","R","99.5","-0.4","A"\n'
    '"2023","CA","10 - 12","T","R","102.0","0.5","A"\n'
    '"2024","B-S","05 - 96","T","R","100.2","0.7","A"\n'
)

_MIGRATION_CSV = (
    '"YEAR","POPULATION_CHANGE_COMPONENT","VALUE","OBS_STATUS"\n'
    '"2023","NMIG","51000","A"\n'
    '"2023","LIVB","85000","A"\n'
    '"2024","NMIG","49000","A"\n'
)

_GDP_CSV = (
    '"PERIOD","INDICATOR","UNIT_MEA","VALUE","OBS_STATUS"\n'
    '"2023","GDP per capita at previous year\'s prices","%","1.2","R"\n'
    '"2023","Labour productivity in actual hours worked","%","0.5","R"\n'
    '"2024","GDP per capita at previous year\'s prices","%","0.8","R"\n'
)

_UNEMPLOYMENT_CSV = (
    '"TIME_PERIOD","GEO","ERWP","ERWL","POP1564","UNIT_MEA","OBS_VALUE","OBS_CONFIDENCE","OBS_STATUS"\n'
    '"2023",CH,Total,1,1,pers in %,4.90,0.11,A\n'
    '"2023",CH011,Total,1,1,pers in %,3.20,0.20,A\n'
    '"2023",CH,Total,1,1,pers,220000,2.1,A\n'
    '"2024",CH,Total,1,1,pers in %,5.14,0.11,A\n'
)


@respx.mock
def test_resolve_asset_url_extracts_current_asset_from_code_page() -> None:
    respx.get("https://www.bfs.admin.ch/asset/de/some-code").mock(
        return_value=Response(
            200,
            text='<a href="https://dam-api.bfs.admin.ch/hub/api/dam/assets/999/master">x</a>',
        )
    )

    assert (
        _resolve_asset_url("some-code")
        == "https://dam-api.bfs.admin.ch/hub/api/dam/assets/999/master"
    )


@respx.mock
def test_resolve_asset_url_raises_when_code_page_has_no_asset_link() -> None:
    respx.get("https://www.bfs.admin.ch/asset/de/dead-code").mock(
        return_value=Response(200, text="<p>Not found</p>")
    )

    with pytest.raises(ValueError, match="dead-code"):
        _resolve_asset_url("dead-code")


@respx.mock
def test_fetch_real_wages_filters_to_whole_economy_real_series() -> None:
    _mock_resolved_asset(_WAGE_INDEX_CODE, "111", _WAGE_CSV.encode())

    observations = fetch_real_wages()

    assert len(observations) == 2
    assert observations[0].date.isoformat() == "2023-01-01"
    assert observations[0].value == 99.5
    assert observations[1].date.isoformat() == "2024-01-01"


@respx.mock
def test_fetch_net_migration_filters_to_nmig_component() -> None:
    _mock_resolved_asset(_POPULATION_COMPONENTS_CODE, "222", _MIGRATION_CSV.encode())

    observations = fetch_net_migration()

    assert len(observations) == 2
    assert observations[0].value == 51000.0
    assert observations[1].value == 49000.0


@respx.mock
def test_fetch_gdp_growth_filters_to_gdp_per_capita_indicator() -> None:
    _mock_resolved_asset(_GDP_GROWTH_CODE, "333", _GDP_CSV.encode())

    observations = fetch_gdp_growth()

    assert len(observations) == 2
    assert observations[0].date.isoformat() == "2023-01-01"
    assert observations[0].value == 1.2
    assert observations[1].value == 0.8


@respx.mock
def test_fetch_unemployment_rate_filters_to_national_percentage_series() -> None:
    _mock_resolved_asset(_UNEMPLOYMENT_CODE, "444", _UNEMPLOYMENT_CSV.encode())

    observations = fetch_unemployment_rate()

    assert len(observations) == 2
    assert observations[0].date.isoformat() == "2023-01-01"
    assert observations[0].value == 4.90
    assert observations[1].value == 5.14


_CPI_INDEX_CSV = (
    '"jahr","monat","index"\n'
    '"2023","Januar",100.0\n'
    '"2023","Februar",101.0\n'
    '"2024","Januar",102.0\n'
    '"2024","Februar",103.02\n'
)


@respx.mock
def test_fetch_cpi_inflation_computes_year_over_year_change() -> None:
    respx.get(_CPI_INDEX_CSV_URL).mock(
        return_value=Response(200, content=_CPI_INDEX_CSV.encode("utf-8-sig"))
    )

    observations = fetch_cpi_inflation()

    assert len(observations) == 2
    assert observations[0].date.isoformat() == "2024-01-01"
    assert round(observations[0].value, 2) == 2.0
    assert observations[1].date.isoformat() == "2024-02-01"
    assert round(observations[1].value, 2) == 2.0


_JOB_VACANCIES_PX = (
    'STUB="Offene Stellen","Grossregion";\n'
    'VALUES("Quartal")="2020Q1","2020Q2";\n'
    "DATA=\n"
    '"..." 10\n'
    "20 30\n"
    "40 50\n"
    "60 70;\n"
)


@respx.mock
def test_fetch_job_vacancies_reads_first_data_row_as_national_total() -> None:
    _mock_resolved_asset(_JOB_VACANCIES_CODE, "555", _JOB_VACANCIES_PX.encode("iso-8859-15"))

    observations = fetch_job_vacancies()

    assert len(observations) == 1
    assert observations[0].date.isoformat() == "2020-04-01"
    assert observations[0].value == 10.0


@respx.mock
def test_fetch_hospital_beds_reads_total_row_per_year_sheet() -> None:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for year, beds in ((2022, 37969.9), (2023, 38100.0)):
        sheet = workbook.create_sheet(str(year))
        sheet.append(["header"])
        sheet.append([None])
        sheet.append(["Total", 22522.1, 1284600, 8541.2, 83876, 6906.6, 94827, beds, 1463303])
        sheet.append(
            ["Genferseeregion", 4686.9, 231622, 1052.4, 12274, 1587.2, 22618, 7326.5, 266514]
        )
    buffer = BytesIO()
    workbook.save(buffer)

    _mock_resolved_asset(_HOSPITAL_BEDS_CODE, "666", buffer.getvalue())

    observations = fetch_hospital_beds()

    assert len(observations) == 2
    assert observations[0].date.isoformat() == "2022-01-01"
    assert observations[0].value == 37969.9
    assert observations[1].date.isoformat() == "2023-01-01"
    assert observations[1].value == 38100.0


@respx.mock
def test_fetch_life_expectancy_averages_men_and_women_by_year() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["header"])
    sheet.append(["filler"])
    sheet.append(["filler"])
    sheet.append(["filler"])
    sheet.append([None, "Männer", None, "Frauen", None])
    sheet.append([None, 2020, 2021, 2020, 2021])
    sheet.append(["Bei Geburt", 81.0, 81.5, 85.0, 85.5])
    buffer = BytesIO()
    workbook.save(buffer)

    _mock_resolved_asset(_LIFE_EXPECTANCY_CODE, "777", buffer.getvalue())

    observations = fetch_life_expectancy()

    assert len(observations) == 2
    assert observations[0].date.isoformat() == "2020-01-01"
    assert observations[0].value == 83.0
    assert observations[1].date.isoformat() == "2021-01-01"
    assert observations[1].value == 83.5


@respx.mock
def test_fetch_co2_emissions_sums_across_sector_blocks_per_year() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Energie"])
    sheet.append([None, 2022, 1.0, 1.0, 1.0, 25.0])
    sheet.append([None, 2023, 1.0, 1.0, 1.0, 24.0])
    sheet.append(["Transport"])
    sheet.append([None, 2022, 1.0, 1.0, 1.0, 15.0])
    sheet.append([None, 2023, 1.0, 1.0, 1.0, 14.0])
    buffer = BytesIO()
    workbook.save(buffer)

    _mock_resolved_asset(_CO2_EMISSIONS_CODE, "888", buffer.getvalue())

    observations = fetch_co2_emissions()

    assert len(observations) == 2
    assert observations[0].date.isoformat() == "2022-01-01"
    assert observations[0].value == 40.0
    assert observations[1].date.isoformat() == "2023-01-01"
    assert observations[1].value == 38.0
