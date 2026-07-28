"""Swiss Federal Statistical Office (BFS) collector.

BFS's PxWeb API endpoint exists but its query endpoints reject every standard
PxWeb request pattern we tried (see docs/DESIGN.md), likely broken or
non-standard on their side. The reliable path is opendata.swiss (CKAN) asset
downloads, which BFS publishes as versioned XLS/CSV snapshots rather than a
live feed. A browser-like User-Agent is required on every request here,
opendata.swiss 403s without one.

Each numeric dam-api asset ID is a permanently frozen document: confirmed by
finding two real cases (CPI, life expectancy) where BFS published a new
edition under a brand-new numeric ID and left the old one abandoned forever
at its original content, rather than updating it in place. BFS does maintain
one genuinely stable pointer though -- each STAT-TAB/PX-table has a permalink
code (e.g. "je-d-01.04.02.03.01") whose landing page keeps linking to
whichever numeric asset is current. `_resolve_asset_url` resolves through
that code every run, so these collectors survive BFS republishing instead of
silently going stale like the original CPI source did.
"""

from __future__ import annotations

import csv
import datetime as dt
import re
from io import BytesIO, StringIO

import httpx
import openpyxl

from ch_at_a_glance.collectors.base import RawObservation, fetch_with_retry

_HEADERS = {"User-Agent": "ch-at-a-glance/0.1 (personal project, non-commercial)"}

# National CPI index level, mirrored by canton Zug's open-data portal. The
# original BFS dam-api asset for this (15964066) turned out to be a frozen
# snapshot that stopped updating in December 2020 -- this source is the same
# national index, but actually kept current (verified: updates through most
# recent published month, not stuck years in the past).
_CPI_INDEX_CSV_URL = "https://data.zg.ch/store/1/resource/334"
_GERMAN_MONTHS = {
    "Januar": 1,
    "Februar": 2,
    "März": 3,
    "April": 4,
    "Mai": 5,
    "Juni": 6,
    "Juli": 7,
    "August": 8,
    "September": 9,
    "Oktober": 10,
    "November": 11,
    "Dezember": 12,
}

# "Schweizerischer Lohnindex, Index und Veraenderung, Basis 2015=100, NOGA08".
_WAGE_INDEX_CODE = "ts-x-03.04.03.00.05"

# "Straf- und Massnahmenvollzug: Einweisungen, mittlerer Bestand, Aufenthaltstage".
_PRISON_CODE = "je-d-19.04.02.02"

# "Komponenten der Entwicklung der staendigen Wohnbevoelkerung, 1861-2024".
_POPULATION_COMPONENTS_CODE = "ts-x-01.02.04.05-a"

# "Zerlegung der Wachstumsrate des BIP pro Kopf" (decomposition of GDP per capita growth).
_GDP_GROWTH_CODE = "ts-x-04.02.01.06"

# "Erwerbs- und Erwerbslosenquote nach Kanton" (economic activity and unemployment rate).
_UNEMPLOYMENT_CODE = "ts-x-40.02.03.02.03"

# "Offene Stellen nach Grossregion" (job vacancies by major region), PC-Axis format.
_JOB_VACANCIES_CODE = "px-x-0602000000_104"

# "Krankenhaeuser: Betten und Hospitalisierungen nach Aktivitaetstyp und Kanton".
_HOSPITAL_BEDS_CODE = "je-d-14.04.01.02"

# "Lebenserwartung" (life expectancy at birth, by sex).
_LIFE_EXPECTANCY_CODE = "je-d-01.04.02.03.01"

# "Treibhausgasemissionen nach Verursachergruppen" (greenhouse gas emissions by source sector).
_CO2_EMISSIONS_CODE = "je-d-02.03.02.03"

_ASSET_LINK_PATTERN = re.compile(r"dam-api\.bfs\.admin\.ch/hub/api/dam/assets/(\d+)")


def _resolve_asset_url(code: str) -> str:
    """Resolve a BFS permalink code to whatever numeric dam-api asset
    currently backs it (see module docstring for why this indirection
    exists instead of a hardcoded asset URL)."""
    with httpx.Client(timeout=30.0, headers=_HEADERS, follow_redirects=True) as client:
        response = fetch_with_retry(client, "GET", f"https://www.bfs.admin.ch/asset/de/{code}")
    match = _ASSET_LINK_PATTERN.search(response.text)
    if match is None:
        raise ValueError(f"Could not resolve a current dam-api asset for BFS code {code!r}")
    return f"https://dam-api.bfs.admin.ch/hub/api/dam/assets/{match.group(1)}/master"


def _fetch_csv_rows(url: str) -> list[dict[str, str]]:
    with httpx.Client(timeout=30.0, headers=_HEADERS, follow_redirects=True) as client:
        response = fetch_with_retry(client, "GET", url)
    text = response.content.decode("utf-8-sig")
    return list(csv.DictReader(StringIO(text)))


def _fetch_csv_rows_by_code(code: str) -> list[dict[str, str]]:
    return _fetch_csv_rows(_resolve_asset_url(code))


def _fetch_workbook_by_code(code: str) -> openpyxl.Workbook:
    with httpx.Client(timeout=30.0, headers=_HEADERS, follow_redirects=True) as client:
        response = fetch_with_retry(client, "GET", _resolve_asset_url(code))
    return openpyxl.load_workbook(BytesIO(response.content), data_only=True)


def fetch_cpi_inflation() -> list[RawObservation]:
    """Swiss CPI, change vs. same month previous year, monthly, national index.

    The CSV publishes the raw index level (not the % change), so this
    computes each month's year-over-year change directly against the same
    calendar month one year earlier.
    """
    rows = _fetch_csv_rows(_CPI_INDEX_CSV_URL)
    index_by_month: dict[tuple[int, int], float] = {}
    for row in rows:
        month = _GERMAN_MONTHS.get(row["monat"])
        if month is None:
            continue
        index_by_month[(int(row["jahr"]), month)] = float(row["index"])

    observations: list[RawObservation] = []
    for (year, month), value in sorted(index_by_month.items()):
        prior_value = index_by_month.get((year - 1, month))
        if prior_value is None:
            continue
        observations.append(
            RawObservation(
                date=dt.date(year, month, 1), value=(value / prior_value - 1) * 100
            )
        )
    return observations


def fetch_real_wages() -> list[RawObservation]:
    """Swiss Wage Index, real (inflation-adjusted), whole economy, annual.

    Filters the CSV to SECTION "B-S" (all NOGA sections, i.e. whole economy),
    SEX "T" (total), WAGE_TYPE "R" (real, as opposed to "N" nominal).
    """
    rows = _fetch_csv_rows_by_code(_WAGE_INDEX_CODE)
    observations: list[RawObservation] = []
    for row in rows:
        if row["SECTION"] != "B-S" or row["SEX"] != "T" or row["WAGE_TYPE"] != "R":
            continue
        observations.append(
            RawObservation(date=dt.date(int(row["YEAR"]), 1, 1), value=float(row["VALUE"]))
        )
    return observations


def fetch_prison_population() -> list[RawObservation]:
    """Average daily prison population ("effectif moyen"), annual.

    Source sheet `DATA`: one row per year, column index 2 is the average
    population; the first three rows are titles/headers, not data.
    """
    workbook = _fetch_workbook_by_code(_PRISON_CODE)
    sheet = workbook["DATA"]
    observations: list[RawObservation] = []
    for row in sheet.iter_rows(values_only=True):
        if not isinstance(row[0], int) or not isinstance(row[2], int | float):
            continue
        observations.append(RawObservation(date=dt.date(row[0], 1, 1), value=float(row[2])))
    return observations


def fetch_net_migration() -> list[RawObservation]:
    """Net migration, annual, from BFS's population change components series."""
    rows = _fetch_csv_rows_by_code(_POPULATION_COMPONENTS_CODE)
    observations: list[RawObservation] = []
    for row in rows:
        if row["POPULATION_CHANGE_COMPONENT"] != "NMIG":
            continue
        observations.append(
            RawObservation(date=dt.date(int(row["YEAR"]), 1, 1), value=float(row["VALUE"]))
        )
    return observations


def fetch_gdp_growth() -> list[RawObservation]:
    """Real GDP per capita growth, annual, from BFS's growth decomposition series."""
    rows = _fetch_csv_rows_by_code(_GDP_GROWTH_CODE)
    observations: list[RawObservation] = []
    for row in rows:
        if row["INDICATOR"] != "GDP per capita at previous year's prices":
            continue
        observations.append(
            RawObservation(date=dt.date(int(row["PERIOD"]), 1, 1), value=float(row["VALUE"]))
        )
    return observations


def fetch_unemployment_rate() -> list[RawObservation]:
    """Unemployment rate, national (ILO/SAKE definition), annual.

    Filters the cantonal-breakdown CSV to GEO "CH" (national aggregate),
    ERWL "1" (unemployed) as a percentage of the working-age population.
    """
    rows = _fetch_csv_rows_by_code(_UNEMPLOYMENT_CODE)
    observations: list[RawObservation] = []
    for row in rows:
        if (
            row["GEO"] != "CH"
            or row["ERWP"] != "Total"
            or row["ERWL"] != "1"
            or row["UNIT_MEA"] != "pers in %"
        ):
            continue
        year = int(row["TIME_PERIOD"].strip('"'))
        observations.append(RawObservation(date=dt.date(year, 1, 1), value=float(row["OBS_VALUE"])))
    return observations


def _px_quoted_list(text: str, marker: str) -> list[str]:
    start = text.index(marker) + len(marker)
    end = text.index(";", start)
    return re.findall(r'"([^"]*)"', text[start:end])


def fetch_job_vacancies() -> list[RawObservation]:
    """Job vacancies, national ("Schweiz"), quarterly, raw count.

    BFS only publishes this as a PC-Axis (.px) file, not CSV/XLS. Rather than
    pull in a PC-Axis parsing library for one indicator, this reads the DATA
    block directly: STUB declares ("Offene Stellen", "Grossregion") in that
    order, so DATA rows are ordered measure-then-region with region cycling
    fastest. The first row (117 quarterly values) is therefore
    "Offene Stellen - Total" for "Schweiz" -- the national count.
    """
    with httpx.Client(timeout=30.0, headers=_HEADERS, follow_redirects=True) as client:
        response = fetch_with_retry(client, "GET", _resolve_asset_url(_JOB_VACANCIES_CODE))
    text = response.content.decode("iso-8859-15")

    quarters = _px_quoted_list(text, 'VALUES("Quartal")=')
    data_start = text.index("DATA=") + len("DATA=")
    tokens = text[data_start:].strip().rstrip(";").split()
    national_row = [token.strip('"') for token in tokens[: len(quarters)]]

    observations: list[RawObservation] = []
    for quarter, token in zip(quarters, national_row, strict=True):
        if token == "...":
            continue
        year, q = int(quarter[:4]), int(quarter[5])
        observations.append(
            RawObservation(date=dt.date(year, (q - 1) * 3 + 1, 1), value=float(token))
        )
    return observations


def fetch_hospital_beds() -> list[RawObservation]:
    """Total hospital beds, national, annual.

    One workbook sheet per year, cantonal breakdown. The "Total" row (right
    after the header block, before the per-region rows) is the national
    figure; column offsets are stable across sheets even though header
    labels vary slightly year to year.
    """
    workbook = _fetch_workbook_by_code(_HOSPITAL_BEDS_CODE)
    observations: list[RawObservation] = []
    for sheet_name in workbook.sheetnames:
        if not sheet_name.isdigit():
            continue
        sheet = workbook[sheet_name]
        total_row = next(
            (row for row in sheet.iter_rows(values_only=True) if row[0] == "Total"), None
        )
        if total_row is None or not isinstance(total_row[7], int | float):
            continue
        observations.append(
            RawObservation(date=dt.date(int(sheet_name), 1, 1), value=float(total_row[7]))
        )
    observations.sort(key=lambda obs: obs.date)
    return observations


def fetch_life_expectancy() -> list[RawObservation]:
    """Life expectancy at birth, national, annual.

    The sheet has an annual block of columns for men followed by an
    identical annual block for women, both on the same "Bei Geburt" (at
    birth) row. The "Maenner"/"Frauen" sex label (row 4) is only set on the
    first column of each block (a merged cell), so it's forward-filled here
    to tell the two blocks apart; national figure is the simple average of
    the two, not population-weighted.
    """
    workbook = _fetch_workbook_by_code(_LIFE_EXPECTANCY_CODE)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    sex_row, year_row = rows[4], rows[5]
    at_birth_row = next(row for row in rows if row[0] == "Bei Geburt")

    current_sex: object = None
    by_year: dict[int, dict[object, float]] = {}
    for sex_label, year, value in zip(sex_row, year_row, at_birth_row, strict=True):
        if sex_label is not None:
            current_sex = sex_label
        if not isinstance(year, int) or not isinstance(value, int | float) or current_sex is None:
            continue
        by_year.setdefault(year, {})[current_sex] = value

    observations: list[RawObservation] = []
    for year, values in sorted(by_year.items()):
        if "Männer" not in values or "Frauen" not in values:
            continue
        men, women = values["Männer"], values["Frauen"]
        observations.append(RawObservation(date=dt.date(year, 1, 1), value=(men + women) / 2))
    return observations


def fetch_co2_emissions() -> list[RawObservation]:
    """Total greenhouse gas emissions, national, annual, CO2-equivalent (Mio. t).

    The sheet has one data block per source sector (energy, transport,
    industry, agriculture, waste), each with its own "Jahr" (year) column.
    There is no precomputed grand-total row, so this sums the CO2-equivalent
    column across every sector for each year.
    """
    workbook = _fetch_workbook_by_code(_CO2_EMISSIONS_CODE)
    sheet = workbook[workbook.sheetnames[0]]
    totals: dict[int, float] = {}
    for row in sheet.iter_rows(values_only=True):
        year, co2_equivalent = row[1], row[5]
        if isinstance(year, int) and isinstance(co2_equivalent, int | float):
            totals[year] = totals.get(year, 0.0) + co2_equivalent
    return [
        RawObservation(date=dt.date(year, 1, 1), value=total)
        for year, total in sorted(totals.items())
    ]
